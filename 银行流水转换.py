#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
银行流水转换工具
将银行对账单 XLS 转换为格式化 Word 文档
不确定或缺失的字段会高亮显示并附注说明
"""

import xlrd
import random
import string
import os
import sys
import time
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from copy import deepcopy

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════
#  常量配置
# ══════════════════════════════════════════════════════════════════

# 高亮颜色（黄色）
HIGHLIGHT_COLOR = "FFFF00"

# 列宽（单位：厘米）
COL_WIDTHS = {
    "序号":   0.8,
    "记账日": 1.5,
    "起息日": 1.5,
    "交易类型": 1.8,
    "凭证号码": 6.5,
    "借方":   2.0,
    "贷方":   2.0,
    "余额":   2.0,
    "机构流水": 3.5,
    "备注":   3.0,
    "问题说明": 3.5,
}

COL_KEYS = list(COL_WIDTHS.keys())

# ══════════════════════════════════════════════════════════════════
#  辅助函数
# ══════════════════════════════════════════════════════════════════

def rnd8():
    return ''.join([str(random.randint(0, 9)) for _ in range(8)])

def rnd13():
    chars = string.ascii_lowercase + string.digits
    return ''.join(random.choices(chars, k=13))

def to_date(d):
    s = str(int(d)) if isinstance(d, float) else str(d)
    return s

def to_display_date(d):
    s = to_date(d)
    return s[2:] if len(s) == 8 else s

def fmt_amount(v):
    return f'{abs(v):,.2f}'

def flow_seq(flow_str):
    s = str(flow_str)
    return s[:9] if len(s) == 12 else s[:8] if len(s) == 11 else s[:9]


# ══════════════════════════════════════════════════════════════════
#  转换规则
# ══════════════════════════════════════════════════════════════════

def transform(row, counters):
    """
    对每一笔交易应用转换规则。
    返回字段字典，以及问题说明列表。
    """
    issues = []  # 记录问题说明
    highlights = set()  # 需要高亮的字段名

    tx   = str(row[0])   # 交易类型
    biz  = str(row[1])   # 业务类型
    payer_code = str(row[2]) if row[2] else ''
    payer_bank = str(row[3]) if row[3] else ''
    payer_name = str(row[5]) if row[5] else ''
    payee_bank = str(row[7]) if row[7] else ''
    payee_name = str(row[9]) if row[9] else ''
    date       = to_date(row[10]) if row[10] else ''
    amount     = float(row[13]) if row[13] else 0.0
    flow       = str(row[17]) if row[17] else ''
    zhaiyao    = str(row[23]) if row[23] else ''
    yongtu     = str(row[24]) if row[24] else ''
    fuyan      = str(row[25]) if row[25] else ''
    beizhu     = str(row[26]) if row[26] else ''
    value_date = to_date(row[15]) if row[15] else date

    key = (tx, biz)

    # ── 来账 ────────────────────────────────────────────────────
    if key in [('来账', '小额普通'), ('来账', '大额支付'), ('来账', '退汇')]:
        if not fuyan:
            issues.append('交易附言为空，凭证号码末段缺失')
            highlights.add('凭证号码')
        voucher = zhaiyao + '/' + fuyan
        ref     = '06293/9880809/' + flow_seq(flow)
        notes   = (payer_name + '/' + payer_bank)[:26]
        if not payer_name:
            issues.append('付款人名称为空')
            highlights.add('备注')
        side = 'credit'

    elif key == ('来账', '网上支付'):
        if not zhaiyao.startswith('IBPS'):
            issues.append(f'摘要不以IBPS开头（实际：{zhaiyao[:10]}），去前4字符结果可能有误')
            highlights.add('凭证号码')
        voucher = zhaiyao[4:] + '/' + fuyan
        ref     = '07289/9880809/' + flow_seq(flow)
        notes   = (payer_name + '/' + payer_bank)[:26]
        side = 'credit'

    elif key == ('来账', '转账收入'):
        if not yongtu:
            issues.append('用途为空，凭证号码前段缺失')
            highlights.add('凭证号码')
        voucher = yongtu + '/' + zhaiyao
        ref     = payer_code + '/9880105/' + flow_seq(flow)
        if not payer_code:
            issues.append('付款人开户行号为空，机构代码缺失')
            highlights.add('机构流水')
        notes   = (payer_name + '/' + payer_bank)[:26]
        side = 'credit'

    elif key == ('来账', '结息'):
        voucher = ''
        ref     = '00557/9770100/' + flow_seq(flow)
        notes   = ''
        side = 'credit'

    # ── 往账 ────────────────────────────────────────────────────
    elif key == ('往账', '转账支出'):
        if not yongtu:
            issues.append('用途为空，凭证号码前段缺失')
            highlights.add('凭证号码')
        voucher = yongtu + '/' + zhaiyao
        ref     = '07289/9880105/' + flow_seq(flow)
        notes   = payee_name
        if not payee_name:
            issues.append('收款人名称为空')
            highlights.add('备注')
        side = 'debit'

    elif key == ('往账', '网上支付'):
        if not yongtu:
            issues.append('用途为空，凭证号码用途段缺失')
            highlights.add('凭证号码')
        rand8 = rnd8()
        voucher = '104100000004' + date + rand8 + '/' + yongtu + '/' + fuyan
        issues.append(f'含随机8位数（{rand8}），需核实实际凭证')
        highlights.add('凭证号码')
        ref   = '07289/9880105/' + flow_seq(flow)
        notes = (payee_name + '/' + payee_bank)[:26]
        side  = 'debit'

    elif key in [('往账', '小额普通'), ('往账', '大额支付')]:
        prefix = beizhu[:4].upper() if beizhu else 'BEPS'
        mid12  = beizhu[26:38] if len(beizhu) >= 38 else ''
        if not mid12:
            issues.append('备注字段过短，无法提取机构代码（第27-38位）')
            highlights.add('凭证号码')
        rand8  = rnd8()
        voucher = prefix + mid12 + '  ' + date + rand8 + '/' + yongtu + '/' + fuyan
        issues.append(f'含随机8位数（{rand8}），需核实实际凭证')
        highlights.add('凭证号码')
        ref   = '07289/9880105/' + flow_seq(flow)
        notes = (payee_name + '/' + payee_bank)[:26]
        side  = 'debit'

    elif key == ('往账', '代发划转'):
        ckey = '代发划转_' + date
        counters[ckey] = counters.get(ckey, 1000) + 1
        counter_str = f'{counters[ckey]:06d}'
        voucher = 'A1589751C1' + date + '5G' + counter_str + '/' + zhaiyao
        ref     = '07289/9880105/' + flow_seq(flow)
        notes   = ''
        side    = 'debit'

    elif key == ('往账', '代收付'):
        if not yongtu:
            issues.append('用途为空，凭证号码前段缺失')
            highlights.add('凭证号码')
        voucher = yongtu + '/' + zhaiyao
        ref     = '07281/9880800/' + flow_seq(flow)
        notes   = payee_name
        side    = 'debit'

    elif key == ('往账', '代收费'):
        rand13  = rnd13()
        voucher = rand13 + date + '/' + zhaiyao
        issues.append(f'含随机13位字符（{rand13}），需核实实际凭证')
        highlights.add('凭证号码')
        ref   = '07289/9880800/' + flow_seq(flow)
        notes = payee_name
        side  = 'debit'

    else:
        # 未知交易类型，使用默认规则
        voucher = (yongtu + '/' if yongtu else '') + zhaiyao
        ref     = '07289/9880105/' + flow_seq(flow)
        notes   = payee_name if tx == '往账' else (payer_name + '/' + payer_bank)[:26]
        side    = 'debit' if tx == '往账' else 'credit'
        issues.append(f'未知交易类型【{tx}-{biz}】，使用默认规则，请人工核查')
        highlights.update(['凭证号码', '机构流水', '备注'])

    # 流水号为空
    if not flow:
        issues.append('交易流水号为空，机构/柜员/流水末段缺失')
        highlights.add('机构流水')

    return {
        'biz':        biz,
        'date':       to_display_date(date),
        'value_date': to_display_date(value_date),
        'amount':     abs(amount),
        'side':       side,
        'voucher':    voucher,
        'ref':        ref,
        'notes':      notes,
        'issues':     issues,
        'highlights': highlights,
    }


# ══════════════════════════════════════════════════════════════════
#  读取 XLS
# ══════════════════════════════════════════════════════════════════

def read_xls(path):
    wb = xlrd.open_workbook(path)
    ws = wb.sheets()[0]

    account    = str(ws.row_values(1)[1])
    time_range = str(ws.row_values(7)[1])
    start_date = time_range[:8]
    end_date   = time_range[-8:]

    raw_txs = []
    for r in range(9, ws.nrows):
        row = ws.row_values(r)
        if not row[0]:
            continue
        while len(row) < 38:
            row.append('')
        raw_txs.append(row)

    # 从第一笔交易推算期初余额
    if raw_txs:
        first = raw_txs[0]
        bal_str = str(first[14]).replace(',', '')
        try:
            opening = float(bal_str) - float(first[13])
        except:
            opening = 0.0
    else:
        opening = 0.0

    return account, start_date, end_date, opening, raw_txs


# ══════════════════════════════════════════════════════════════════
#  生成 Excel 文档
# ══════════════════════════════════════════════════════════════════

# 列宽（Excel 字符单位）
XL_COL_WIDTHS = {
    '序号': 6, '记账日': 10, '起息日': 10, '交易类型': 10,
    '凭证号码': 55, '借方': 16, '贷方': 16, '余额': 16,
    '机构流水': 28, '备注': 26, '问题说明': 32,
}

FILL_HEADER  = PatternFill('solid', fgColor='D0D8E8')
FILL_YELLOW  = PatternFill('solid', fgColor='FFFF00')
FILL_ROW_ODD = PatternFill('solid', fgColor='FFFFFF')
FILL_ROW_EVN = PatternFill('solid', fgColor='F0F4FF')
FILL_SUMM    = PatternFill('solid', fgColor='E8EDF5')

FONT_HDR  = Font(name='微软雅黑', size=9, bold=True)
FONT_DATA = Font(name='微软雅黑', size=9)
FONT_WARN = Font(name='微软雅黑', size=9, color='CC0000')

ALIGN_CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LFT = Alignment(horizontal='left',   vertical='center', wrap_text=True)

_side = Side(style='thin', color='AAAAAA')
BORDER = Border(left=_side, right=_side, top=_side, bottom=_side)


def build_excel(tx_data, output_path, account, start_date, end_date, opening_balance):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '交易明细'

    # ── 标题行 ───────────────────────────────────────────────────
    ws.merge_cells('A1:K1')
    ws['A1'] = '银行账户交易明细'
    ws['A1'].font = Font(name='微软雅黑', size=13, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A2:K2')
    ws['A2'] = f'账号：{account}　　期间：{start_date} — {end_date}　　期初余额：{fmt_amount(opening_balance)}'
    ws['A2'].font = Font(name='微软雅黑', size=9)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16

    # ── 表头（第3行）───────────────────────────────────────────
    HDR_LABELS = {
        '序号':   '序号\nNo.',
        '记账日': '记账日\nBk.D.',
        '起息日': '起息日\nVal.d.',
        '交易类型': '交易类型\nType',
        '凭证号码': '凭证号码/业务编号/用途/摘要\nVou. No./Trans. No./Details',
        '借方':   '借方发生额\nDebit Amount',
        '贷方':   '贷方发生额\nCredit Amount',
        '余额':   '余额\nBalance',
        '机构流水': '机构/柜员/流水\nReference No.',
        '备注':   '备注\nNotes',
        '问题说明': '⚠ 问题说明\nIssue Notes',
    }
    for ci, key in enumerate(COL_KEYS, 1):
        cell = ws.cell(row=3, column=ci, value=HDR_LABELS[key])
        cell.font      = FONT_HDR
        cell.fill      = FILL_HEADER
        cell.alignment = ALIGN_CTR
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = XL_COL_WIDTHS[key]
    ws.row_dimensions[3].height = 30

    # ── 数据行 ──────────────────────────────────────────────────
    balance = opening_balance
    problem_count = 0

    for seq, tx in enumerate(tx_data, 1):
        r = seq + 3   # row index in sheet
        if tx['side'] == 'debit':
            balance -= tx['amount']
        else:
            balance += tx['amount']

        hl          = tx['highlights']
        issues_text = '\n'.join(tx['issues']) if tx['issues'] else ''
        has_issue   = bool(tx['issues'])
        if has_issue:
            problem_count += 1

        row_fill = FILL_ROW_EVN if seq % 2 == 0 else FILL_ROW_ODD

        values = [
            str(seq),
            tx['date'],
            tx['value_date'] if tx['value_date'] != tx['date'] else '',
            tx['biz'],
            tx['voucher'],
            fmt_amount(tx['amount']) if tx['side'] == 'debit'  else '',
            fmt_amount(tx['amount']) if tx['side'] == 'credit' else '',
            fmt_amount(balance),
            tx['ref'],
            tx['notes'],
            issues_text,
        ]

        for ci, (key, val) in enumerate(zip(COL_KEYS, values), 1):
            cell = ws.cell(row=r, column=ci, value=val)
            do_hl = (key in hl) or (key == '问题说明' and has_issue)
            cell.fill      = FILL_YELLOW if do_hl else row_fill
            cell.font      = FONT_WARN if (key == '问题说明' and has_issue) else FONT_DATA
            cell.alignment = ALIGN_LFT if key in ('凭证号码', '备注', '问题说明') else ALIGN_CTR
            cell.border    = BORDER

        ws.row_dimensions[r].height = 15

    # ── 汇总行 ──────────────────────────────────────────────────
    total_debit  = sum(t['amount'] for t in tx_data if t['side'] == 'debit')
    total_credit = sum(t['amount'] for t in tx_data if t['side'] == 'credit')
    sr = len(tx_data) + 4
    summ_vals = ['合计', '', '', '', f'共 {len(tx_data)} 笔　⚠ 问题项 {problem_count} 笔',
                 fmt_amount(total_debit), fmt_amount(total_credit), fmt_amount(balance),
                 '', '', '']
    for ci, val in enumerate(summ_vals, 1):
        cell = ws.cell(row=sr, column=ci, value=val)
        cell.font      = Font(name='微软雅黑', size=9, bold=True)
        cell.fill      = FILL_SUMM
        cell.alignment = ALIGN_CTR
        cell.border    = BORDER
    ws.row_dimensions[sr].height = 16

    # 冻结表头
    ws.freeze_panes = 'A4'

    wb.save(output_path)
    return problem_count


# ══════════════════════════════════════════════════════════════════
#  GUI 界面
# ══════════════════════════════════════════════════════════════════

class App:
    def __init__(self, root):
        self.root = root
        root.title('银行流水转换工具')
        root.resizable(True, True)

        pad = dict(padx=10, pady=6)

        # 输入文件
        tk.Label(root, text='输入文件（XLS）：', anchor='w').grid(row=0, column=0, sticky='w', **pad)
        self.xls_var = tk.StringVar()
        tk.Entry(root, textvariable=self.xls_var, width=50).grid(row=0, column=1, **pad)
        tk.Button(root, text='选择文件', command=self.pick_xls).grid(row=0, column=2, **pad)

        # 输出文件
        tk.Label(root, text='输出文件（XLSX）：', anchor='w').grid(row=1, column=0, sticky='w', **pad)
        self.out_var = tk.StringVar()
        tk.Entry(root, textvariable=self.out_var, width=50).grid(row=1, column=1, **pad)
        tk.Button(root, text='选择路径', command=self.pick_out).grid(row=1, column=2, **pad)

        # 随机种子选项
        self.seed_var = tk.BooleanVar(value=True)
        tk.Checkbutton(root, text='固定随机数（确保每次生成结果一致）', variable=self.seed_var).grid(
            row=2, column=0, columnspan=3, sticky='w', padx=10)

        # 转换按钮
        tk.Button(root, text='开始转换', command=self.run,
                  bg='#336699', fg='white', font=('', 11, 'bold'),
                  width=20, height=2).grid(row=3, column=0, columnspan=3, pady=12)

        # 状态栏
        self.status_var = tk.StringVar(value='请选择输入文件后点击"开始转换"')
        tk.Label(root, textvariable=self.status_var, fg='#555', anchor='w').grid(
            row=4, column=0, columnspan=3, sticky='w', padx=10, pady=4)

        # 进度条（确定模式，显示实际进度）
        self.progress = ttk.Progressbar(root, length=500, mode='determinate')
        self.progress.grid(row=5, column=0, columnspan=3, padx=10, pady=4)

        # 日志区域
        log_frame = tk.Frame(root)
        log_frame.grid(row=6, column=0, columnspan=3, padx=10, pady=(0, 10), sticky='nsew')
        root.grid_rowconfigure(6, weight=1)
        root.grid_columnconfigure(1, weight=1)

        tk.Label(log_frame, text='处理日志：', anchor='w').pack(anchor='w')
        self.log_text = tk.Text(log_frame, height=14, width=78, state='disabled',
                                bg='#1e1e1e', fg='#d4d4d4', font=('Courier', 9),
                                relief='flat', bd=1)
        scrollbar = tk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # 颜色标签
        self.log_text.tag_configure('ok',    foreground='#6fcf97')
        self.log_text.tag_configure('warn',  foreground='#f2c94c')
        self.log_text.tag_configure('error', foreground='#eb5757')
        self.log_text.tag_configure('info',  foreground='#56cfe1')
        self.log_text.tag_configure('dim',   foreground='#888888')

    def log(self, msg, tag='', flush=False):
        """Thread-safe log: schedules UI update on main thread via after()."""
        ts = time.strftime('%H:%M:%S')
        line = f'[{ts}] {msg}\n'
        def _insert():
            self.log_text.configure(state='normal')
            self.log_text.insert('end', line, tag)
            self.log_text.see('end')
            self.log_text.configure(state='disabled')
        self.root.after(0, _insert)

    def set_progress(self, value, status=''):
        def _update():
            self.progress['value'] = value
            if status:
                self.status_var.set(status)
        self.root.after(0, _update)

    def pick_xls(self):
        path = filedialog.askopenfilename(
            title='选择银行流水 XLS 文件',
            filetypes=[('Excel 文件', '*.xls *.xlsx'), ('所有文件', '*.*')]
        )
        if path:
            self.xls_var.set(path)
            default_out = os.path.splitext(path)[0] + '_转换结果.xlsx'
            self.out_var.set(default_out)

    def pick_out(self):
        path = filedialog.asksaveasfilename(
            title='保存输出 Excel 文件',
            defaultextension='.xlsx',
            filetypes=[('Excel 文件', '*.xlsx'), ('所有文件', '*.*')]
        )
        if path:
            self.out_var.set(path)

    def run(self):
        xls_path = self.xls_var.get().strip()
        out_path = self.out_var.get().strip()

        if not xls_path:
            messagebox.showerror('错误', '请先选择输入的 XLS 文件')
            return
        if not out_path:
            messagebox.showerror('错误', '请先指定输出文件路径')
            return
        if not os.path.exists(xls_path):
            messagebox.showerror('错误', f'文件不存在：\n{xls_path}')
            return

        if self.seed_var.get():
            random.seed(42)

        # 清空日志
        self.log_text.configure(state='normal')
        self.log_text.delete('1.0', 'end')
        self.log_text.configure(state='disabled')
        self.progress['value'] = 0
        self.status_var.set('正在处理...')

        # 在后台线程运行，保持UI响应
        t = threading.Thread(target=self._run_worker, args=(xls_path, out_path), daemon=True)
        t.start()

    def _run_worker(self, xls_path, out_path):
        t_start = time.time()
        try:
            self.log(f'开始处理：{os.path.basename(xls_path)}', 'info')
            self.set_progress(0, '正在读取文件...')

            # ── 第一步：读取 XLS ──────────────────────────────────
            t0 = time.time()
            account, start_date, end_date, opening, raw_txs = read_xls(xls_path)
            self.log(f'读取完成：{len(raw_txs)} 笔交易，账号 {account}，期间 {start_date}—{end_date}  ({time.time()-t0:.2f}s)', 'ok')
            self.log(f'期初余额：{fmt_amount(opening)}', 'dim')
            self.set_progress(10)

            # ── 第二步：逐笔转换 ──────────────────────────────────
            counters = {}
            tx_data = []
            warn_count = 0
            t0 = time.time()

            for i, row in enumerate(raw_txs):
                result = transform(row, counters)
                tx_data.append(result)

                tx_label = f'第{i+1:>4}笔  {result["date"]}  {result["biz"]:<6}'
                if result['issues']:
                    warn_count += 1
                    for iss in result['issues']:
                        self.log(f'⚠  {tx_label} → {iss}', 'warn')
                else:
                    self.log(f'✓  {tx_label}  {fmt_amount(result["amount"])} {"借" if result["side"]=="debit" else "贷"}', 'dim')

                pct = 10 + int(70 * (i + 1) / len(raw_txs))
                self.set_progress(pct, f'正在转换... {i+1}/{len(raw_txs)} 笔')

            elapsed_tx = time.time() - t0
            self.log(f'转换完成：{len(tx_data)} 笔，耗时 {elapsed_tx:.2f}s，问题 {warn_count} 笔', 'ok')
            self.set_progress(80, '正在生成 Excel 文档...')

            # ── 第三步：生成 Excel ────────────────────────────────
            self.log('开始写入 Excel 文档...', 'info')
            t0 = time.time()
            problems = build_excel(tx_data, out_path, account, start_date, end_date, opening)
            elapsed_word = time.time() - t0

            total_elapsed = time.time() - t_start
            self.log(f'Excel 生成完成，耗时 {elapsed_word:.2f}s', 'ok')
            self.log('─────────────────────────────────────────────────', 'dim')
            self.log(f'全部完成！总耗时 {total_elapsed:.2f}s', 'ok')
            self.log(f'输出文件：{out_path}', 'info')
            self.set_progress(100, f'✅ 完成：{len(tx_data)} 笔，{problems} 笔需核查，总耗时 {total_elapsed:.2f}s')

            self.root.after(0, lambda: messagebox.showinfo('转换完成',
                f'转换完成！\n\n'
                f'共处理 {len(tx_data)} 笔交易\n'
                f'其中 ⚠ {problems} 笔存在问题（已高亮标注）\n'
                f'总耗时：{total_elapsed:.2f} 秒\n\n'
                f'输出文件：\n{out_path}'))

        except Exception as e:
            import traceback
            self.log(f'❌ 错误：{e}', 'error')
            self.log(traceback.format_exc(), 'error')
            self.set_progress(0, f'❌ 错误：{e}')
            self.root.after(0, lambda: messagebox.showerror('转换失败', f'发生错误：\n\n{e}'))


# ══════════════════════════════════════════════════════════════════
#  入口
# ══════════════════════════════════════════════════════════════════

def main():
    root = tk.Tk()
    app = App(root)
    root.mainloop()


if __name__ == '__main__':
    main()
